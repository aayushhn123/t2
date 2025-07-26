import streamlit as st
import pandas as pd
from datetime import datetime, timedelta
from fpdf import FPDF
import os
import re
import io
from PyPDF2 import PdfReader, PdfWriter
from collections import defaultdict
import numpy as np
from openpyxl import load_workbook

# Set page configuration
st.set_page_config(
    page_title="Timetable Optimizer",
    page_icon="🔧",
    layout="wide",
    initial_sidebar_state="expanded"
)

# Custom CSS (reuse from reference file)
st.markdown("""
<style>
    /* Base styles */
    .main-header {
        padding: 2rem;
        border-radius: 10px;
        margin-bottom: 2rem;
        background: linear-gradient(90deg, #951C1C, #C73E1D);
    }
    .main-header h1 {
        color: white;
        text-align: center;
        margin: 0;
        font-size: 2.5rem;
        text-shadow: 2px 2px 4px rgba(0,0,0,0.3);
    }
    .main-header p {
        color: #FFF;
        text-align: center;
        margin: 0.5rem 0 0 0;
        font-size: 1.2rem;
        opacity: 0.9;
    }
    .optimization-card {
        background: #f8f9fa;
        padding: 1.5rem;
        border-radius: 10px;
        margin: 1rem 0;
        border-left: 4px solid #28a745;
    }
    .metric-card {
        background: linear-gradient(135deg, #667eea 0%, #764ba2 100%);
        padding: 1.5rem;
        border-radius: 8px;
        color: white;
        text-align: center;
        margin: 0.5rem;
    }
    .status-success {
        background: #d4edda;
        color: #155724;
        padding: 1rem;
        border-radius: 5px;
        border-left: 4px solid #28a745;
    }
    .status-error {
        background: #f8d7da;
        color: #721c24;
        padding: 1rem;
        border-radius: 5px;
        border-left: 4px solid #dc3545;
    }
    .status-info {
        background: #d1ecf1;
        color: #0c5460;
        padding: 1rem;
        border-radius: 5px;
        border-left: 4px solid #17a2b8;
    }
</style>
""", unsafe_allow_html=True)

# Constants from reference file
BRANCH_FULL_FORM = {
    "B TECH": "BACHELOR OF TECHNOLOGY",
    "B TECH INTG": "BACHELOR OF TECHNOLOGY SIX YEAR INTEGRATED PROGRAM",
    "M TECH": "MASTER OF TECHNOLOGY",
    "MBA TECH": "MASTER OF BUSINESS ADMINISTRATION IN TECHNOLOGY MANAGEMENT",
    "MCA": "MASTER OF COMPUTER APPLICATIONS"
}

LOGO_PATH = "logo.png"

class TimetableOptimizer:
    def __init__(self):
        self.timetable_data = {}
        self.optimization_log = []
        self.holidays = set()
        
    def read_existing_timetable(self, file_path):
        """Read existing timetable from Excel file"""
        st.info("📖 Reading existing timetable...")
        
        # Read all sheets
        xl_file = pd.ExcelFile(file_path)
        sheet_names = xl_file.sheet_names
        
        timetable_data = {}
        
        for sheet_name in sheet_names:
            # Read sheet with multi-index (Date and Time Slot)
            df = pd.read_excel(file_path, sheet_name=sheet_name, index_col=[0, 1])
            
            # Parse sheet name to get branch and semester info
            parts = sheet_name.split('_Sem_')
            if len(parts) >= 2:
                main_branch = parts[0]
                semester_str = parts[1].replace('_Electives', '')
                
                # Convert Roman to integer
                semester = self.roman_to_int(semester_str)
                is_elective = '_Electives' in sheet_name
                
                timetable_data[sheet_name] = {
                    'data': df,
                    'main_branch': main_branch,
                    'semester': semester,
                    'is_elective': is_elective,
                    'sheet_name': sheet_name
                }
        
        self.timetable_data = timetable_data
        st.success(f"✅ Successfully read {len(timetable_data)} sheets")
        return timetable_data
    
    def roman_to_int(self, roman):
        """Convert Roman numeral to integer"""
        roman_values = {
            'I': 1, 'V': 5, 'X': 10, 'L': 50,
            'C': 100, 'D': 500, 'M': 1000
        }
        total = 0
        prev_value = 0
        
        for char in reversed(roman):
            value = roman_values.get(char, 0)
            if value < prev_value:
                total -= value
            else:
                total += value
            prev_value = value
        
        return total
    
    def int_to_roman(self, num):
        """Convert integer to Roman numeral"""
        roman_values = [
            (1000, "M"), (900, "CM"), (500, "D"), (400, "CD"),
            (100, "C"), (90, "XC"), (50, "L"), (40, "XL"),
            (10, "X"), (9, "IX"), (5, "V"), (4, "IV"), (1, "I")
        ]
        result = ""
        for value, numeral in roman_values:
            while num >= value:
                result += numeral
                num -= value
        return result
    
    def analyze_timetable(self):
        """Analyze current timetable to find empty slots and calculate span"""
        analysis = {
            'total_empty_slots': 0,
            'empty_slots_by_date': defaultdict(int),
            'empty_slots_by_sheet': defaultdict(int),
            'current_span': 0,
            'all_dates': set(),
            'exam_dates': set(),
            'oe_dates': set(),
            'oe_info': {}
        }
        
        for sheet_name, sheet_data in self.timetable_data.items():
            df = sheet_data['data']
            
            for (date, time_slot), row in df.iterrows():
                # Parse date
                try:
                    date_obj = pd.to_datetime(date, format="%d-%m-%Y")
                    analysis['all_dates'].add(date_obj)
                except:
                    continue
                
                # Track OE dates specifically
                if sheet_data['is_elective']:
                    has_oe_exam = False
                    for col in df.columns:
                        if not (pd.isna(row[col]) or str(row[col]).strip() == "---"):
                            has_oe_exam = True
                            break
                    if has_oe_exam:
                        analysis['oe_dates'].add(date_obj)
                        if date not in analysis['oe_info']:
                            analysis['oe_info'][date] = []
                        analysis['oe_info'][date].append((time_slot, sheet_name))
                
                # Count empty slots
                for col in df.columns:
                    if pd.isna(row[col]) or str(row[col]).strip() == "---":
                        analysis['total_empty_slots'] += 1
                        analysis['empty_slots_by_date'][date] += 1
                        analysis['empty_slots_by_sheet'][sheet_name] += 1
                    else:
                        analysis['exam_dates'].add(date_obj)
        
        # Calculate current span
        if analysis['exam_dates']:
            min_date = min(analysis['exam_dates'])
            max_date = max(analysis['exam_dates'])
            analysis['current_span'] = (max_date - min_date).days + 1
            analysis['start_date'] = min_date
            analysis['end_date'] = max_date
        
        return analysis
    
    def optimize_timetable(self, holidays=None):
        """Optimize timetable by filling empty slots"""
        if holidays:
            self.holidays = holidays
        
        st.info("🔧 Starting optimization process...")
        
        # Analyze current state
        initial_analysis = self.analyze_timetable()
        
        # Create optimization plan
        optimization_moves = []
        
        # Step 1: Process non-elective sheets first
        st.info("📚 Step 1: Optimizing regular subjects...")
        for sheet_name, sheet_data in self.timetable_data.items():
            if sheet_data['is_elective']:
                continue  # Skip elective sheets for now
            
            df = sheet_data['data'].copy()
            
            # Get all dates and sort them
            all_dates = []
            for (date, time_slot), _ in df.iterrows():
                try:
                    date_obj = pd.to_datetime(date, format="%d-%m-%Y")
                    all_dates.append((date_obj, date, time_slot))
                except:
                    continue
            
            all_dates.sort(key=lambda x: x[0])
            
            # Find movable exams (from later dates to earlier empty slots)
            for i in range(len(all_dates) - 1, 0, -1):
                late_date_obj, late_date, late_time = all_dates[i]
                
                # Skip if date is a holiday
                if late_date_obj.date() in self.holidays:
                    continue
                
                # Check each branch/column
                for col in df.columns:
                    exam = df.loc[(late_date, late_time), col]
                    
                    # Skip if empty
                    if pd.isna(exam) or str(exam).strip() == "---":
                        continue
                    
                    # Try to find earlier empty slot
                    for j in range(i):
                        early_date_obj, early_date, early_time = all_dates[j]
                        
                        # Skip if date is a holiday
                        if early_date_obj.date() in self.holidays:
                            continue
                        
                        # Check if slot is empty
                        current_value = df.loc[(early_date, early_time), col]
                        if pd.isna(current_value) or str(current_value).strip() == "---":
                            # Found empty slot - plan the move
                            optimization_moves.append({
                                'sheet': sheet_name,
                                'exam': exam,
                                'branch': col,
                                'from_date': late_date,
                                'from_time': late_time,
                                'to_date': early_date,
                                'to_time': early_time,
                                'days_saved': (late_date_obj - early_date_obj).days
                            })
                            
                            # Execute the move
                            df.loc[(early_date, early_time), col] = exam
                            df.loc[(late_date, late_time), col] = "---"
                            
                            self.optimization_log.append(
                                f"Moved {exam} from {late_date} to {early_date} (saved {(late_date_obj - early_date_obj).days} days)"
                            )
                            break
            
            # Update the sheet data
            sheet_data['data'] = df
        
        # Step 2: Optimize OE (Open Elective) subjects
        st.info("🎯 Step 2: Optimizing Open Electives...")
        self.optimize_oe_subjects()
        
        # Analyze optimized state
        final_analysis = self.analyze_timetable()
        
        # Calculate improvement
        days_saved = initial_analysis['current_span'] - final_analysis['current_span']
        slots_filled = initial_analysis['total_empty_slots'] - final_analysis['total_empty_slots']
        
        st.success(f"✅ Optimization complete!")
        st.info(f"📊 Days saved: {days_saved} | Slots filled: {slots_filled}")
        
        return {
            'initial_analysis': initial_analysis,
            'final_analysis': final_analysis,
            'optimization_moves': optimization_moves,
            'days_saved': days_saved,
            'slots_filled': slots_filled
        }
    
    def optimize_oe_subjects(self):
        """Optimize OE subjects by finding common earlier slots across all semesters"""
        # First, identify all OE sheets and their current dates
        oe_sheets = {}
        oe_dates = set()
        
        for sheet_name, sheet_data in self.timetable_data.items():
            if sheet_data['is_elective']:
                oe_sheets[sheet_name] = sheet_data
                df = sheet_data['data']
                
                # Get all dates from this OE sheet
                for (date, time_slot), _ in df.iterrows():
                    try:
                        date_obj = pd.to_datetime(date, format="%d-%m-%Y")
                        oe_dates.add((date_obj, date, time_slot))
                    except:
                        continue
        
        if not oe_sheets:
            st.info("No OE sheets found to optimize")
            return
        
        # Sort OE dates
        oe_dates_list = sorted(list(oe_dates), key=lambda x: x[0])
        
        # Find common available slots across ALL sheets (including non-elective)
        common_available_slots = self.find_common_available_slots()
        
        # Try to move OE exams to earlier common slots
        oe_moves = 0
        
        for oe_date_obj, oe_date, oe_time in reversed(oe_dates_list):
            # Check if we can move this OE to an earlier slot
            for slot_date_obj, slot_date, slot_time in common_available_slots:
                if slot_date_obj < oe_date_obj:
                    # This is an earlier slot - check if it's truly available for all OE subjects
                    can_move = True
                    
                    # Verify slot is empty across all relevant sheets
                    for sheet_name, sheet_data in self.timetable_data.items():
                        df = sheet_data['data']
                        if (slot_date, slot_time) in df.index:
                            # Check if this slot has any exams
                            row = df.loc[(slot_date, slot_time)]
                            for col in df.columns:
                                if not (pd.isna(row[col]) or str(row[col]).strip() == "---"):
                                    can_move = False
                                    break
                        if not can_move:
                            break
                    
                    if can_move:
                        # Move all OE exams from the old date to the new date
                        st.info(f"Moving OE exams from {oe_date} to {slot_date}")
                        
                        for oe_sheet_name, oe_sheet_data in oe_sheets.items():
                            df = oe_sheet_data['data']
                            
                            if (oe_date, oe_time) in df.index:
                                # Get the row data
                                old_row = df.loc[(oe_date, oe_time)].copy()
                                
                                # Add new row if it doesn't exist
                                if (slot_date, slot_time) not in df.index:
                                    df.loc[(slot_date, slot_time)] = "---"
                                
                                # Move the data
                                df.loc[(slot_date, slot_time)] = old_row
                                df.loc[(oe_date, oe_time)] = "---"
                                
                                self.optimization_log.append(
                                    f"Moved OE exams from {oe_date} {oe_time} to {slot_date} {slot_time} (saved {(oe_date_obj - slot_date_obj).days} days)"
                                )
                        
                        oe_moves += 1
                        # Remove this slot from available slots
                        common_available_slots.remove((slot_date_obj, slot_date, slot_time))
                        break
        
        st.success(f"✅ Optimized {oe_moves} OE exam slots")
    
    def find_common_available_slots(self):
        """Find slots that are available across all sheets"""
        # Get all possible dates from the timetable
        all_possible_dates = set()
        
        for sheet_name, sheet_data in self.timetable_data.items():
            df = sheet_data['data']
            for (date, time_slot), _ in df.iterrows():
                try:
                    date_obj = pd.to_datetime(date, format="%d-%m-%Y")
                    if date_obj.date() not in self.holidays:
                        all_possible_dates.add((date_obj, date, time_slot))
                except:
                    continue
        
        # Sort dates
        sorted_dates = sorted(list(all_possible_dates), key=lambda x: x[0])
        
        # Find truly empty slots across all sheets
        common_empty_slots = []
        
        for date_obj, date, time_slot in sorted_dates:
            is_empty_everywhere = True
            
            # Check this slot in all sheets
            for sheet_name, sheet_data in self.timetable_data.items():
                df = sheet_data['data']
                
                if (date, time_slot) in df.index:
                    row = df.loc[(date, time_slot)]
                    # Check if any column has an exam
                    for col in df.columns:
                        if not (pd.isna(row[col]) or str(row[col]).strip() == "---"):
                            is_empty_everywhere = False
                            break
                
                if not is_empty_everywhere:
                    break
            
            if is_empty_everywhere:
                common_empty_slots.append((date_obj, date, time_slot))
        
        return common_empty_slots
    
    def save_optimized_excel(self):
        """Save optimized timetable to Excel"""
        output = io.BytesIO()
        
        with pd.ExcelWriter(output, engine='openpyxl') as writer:
            for sheet_name, sheet_data in self.timetable_data.items():
                df = sheet_data['data']
                df.to_excel(writer, sheet_name=sheet_name)
        
        output.seek(0)
        return output
    
    def generate_optimized_pdf(self, output_path):
        """Generate PDF from optimized timetable"""
        # First save to temporary Excel
        temp_excel = "temp_optimized.xlsx"
        excel_data = self.save_optimized_excel()
        
        with open(temp_excel, 'wb') as f:
            f.write(excel_data.getvalue())
        
        # Use the convert_excel_to_pdf function from reference
        convert_excel_to_pdf(temp_excel, output_path)
        
        # Clean up
        if os.path.exists(temp_excel):
            os.remove(temp_excel)

# Include necessary functions from reference file
def wrap_text(pdf, text, col_width):
    """Text wrapping function from reference"""
    words = text.split()
    lines = []
    current_line = ""
    for word in words:
        test_line = word if not current_line else current_line + " " + word
        if pdf.get_string_width(test_line) <= col_width:
            current_line = test_line
        else:
            lines.append(current_line)
            current_line = word
    if current_line:
        lines.append(current_line)
    return lines

def print_row_custom(pdf, row_data, col_widths, line_height=5, header=False):
    """Custom row printing from reference"""
    cell_padding = 2
    header_bg_color = (149, 33, 28)
    header_text_color = (255, 255, 255)
    alt_row_color = (240, 240, 240)

    row_number = getattr(pdf, '_row_counter', 0)
    if header:
        pdf.set_font("Arial", 'B', 10)
        pdf.set_text_color(*header_text_color)
        pdf.set_fill_color(*header_bg_color)
    else:
        pdf.set_font("Arial", size=10)
        pdf.set_text_color(0, 0, 0)
        pdf.set_fill_color(*alt_row_color if row_number % 2 == 1 else (255, 255, 255))

    wrapped_cells = []
    max_lines = 0
    for i, cell_text in enumerate(row_data):
        text = str(cell_text) if cell_text is not None else ""
        avail_w = col_widths[i] - 2 * cell_padding
        lines = wrap_text(pdf, text, avail_w)
        wrapped_cells.append(lines)
        max_lines = max(max_lines, len(lines))

    row_h = line_height * max_lines
    x0, y0 = pdf.get_x(), pdf.get_y()
    if header or row_number % 2 == 1:
        pdf.rect(x0, y0, sum(col_widths), row_h, 'F')

    for i, lines in enumerate(wrapped_cells):
        cx = pdf.get_x()
        pad_v = (row_h - len(lines) * line_height) / 2 if len(lines) < max_lines else 0
        for j, ln in enumerate(lines):
            pdf.set_xy(cx + cell_padding, y0 + j * line_height + pad_v)
            pdf.cell(col_widths[i] - 2 * cell_padding, line_height, ln, border=0, align='C')
        pdf.rect(cx, y0, col_widths[i], row_h)
        pdf.set_xy(cx + col_widths[i], y0)

    setattr(pdf, '_row_counter', row_number + 1)
    pdf.set_xy(x0, y0 + row_h)

def convert_excel_to_pdf(excel_path, pdf_path, sub_branch_cols_per_page=4):
    """Convert Excel to PDF - simplified version from reference"""
    pdf = FPDF(orientation='L', unit='mm', format=(210, 500))
    pdf.set_auto_page_break(auto=False, margin=15)
    pdf.alias_nb_pages()
    
    df_dict = pd.read_excel(excel_path, sheet_name=None, index_col=[0, 1])
    
    for sheet_name, pivot_df in df_dict.items():
        if pivot_df.empty:
            continue
            
        # Parse sheet name
        parts = sheet_name.split('_Sem_')
        main_branch = parts[0]
        main_branch_full = BRANCH_FULL_FORM.get(main_branch, main_branch)
        
        # Add page and content
        pdf.add_page()
        
        # Simple header
        pdf.set_font("Arial", 'B', 16)
        pdf.cell(0, 10, f"{main_branch_full} - {sheet_name}", 0, 1, 'C')
        pdf.ln(10)
        
        # Convert dataframe to table
        pdf.set_font("Arial", size=10)
        
        # Print the table content
        pivot_df_reset = pivot_df.reset_index()
        
        # Header
        headers = list(pivot_df_reset.columns)
        col_width = (pdf.w - 20) / len(headers)
        col_widths = [col_width] * len(headers)
        
        print_row_custom(pdf, headers, col_widths, header=True)
        
        # Data rows
        for idx in range(len(pivot_df_reset)):
            row_data = [str(pivot_df_reset.iloc[idx][col]) for col in headers]
            print_row_custom(pdf, row_data, col_widths)
    
    pdf.output(pdf_path)

def main():
    st.markdown("""
    <div class="main-header">
        <h1>🔧 Exam Timetable Optimizer</h1>
        <p>MUKESH PATEL SCHOOL OF TECHNOLOGY MANAGEMENT & ENGINEERING</p>
        <p style="font-size: 0.9em; opacity: 0.8;">Optimize existing timetables by filling empty slots</p>
    </div>
    """, unsafe_allow_html=True)
    
    # Initialize optimizer
    if 'optimizer' not in st.session_state:
        st.session_state.optimizer = TimetableOptimizer()
    
    optimizer = st.session_state.optimizer
    
    # Sidebar configuration
    with st.sidebar:
        st.markdown("### ⚙️ Configuration")
        
        # Holiday configuration
        st.markdown("#### 📅 Holiday Settings")
        holidays = set()
        
        col1, col2 = st.columns(2)
        with col1:
            if st.checkbox("April 14, 2025", value=True):
                holidays.add(datetime(2025, 4, 14).date())
        with col2:
            if st.checkbox("May 1, 2025", value=True):
                holidays.add(datetime(2025, 5, 1).date())
        
        if st.checkbox("August 15, 2025", value=True):
            holidays.add(datetime(2025, 8, 15).date())
        
        st.markdown("#### 🎯 Optimization Settings")
        max_moves = st.slider("Maximum moves per sheet", 10, 100, 50)
        prioritize_early_slots = st.checkbox("Prioritize earliest slots", value=True)
    
    # Main content
    col1, col2 = st.columns([2, 1])
    
    with col1:
        st.markdown("""
        <div class="optimization-card">
            <h3>📁 Upload Existing Timetable</h3>
            <p>Upload your existing timetable Excel file with empty slots marked as "---"</p>
        </div>
        """, unsafe_allow_html=True)
        
        uploaded_file = st.file_uploader(
            "Choose existing timetable Excel file",
            type=['xlsx', 'xls'],
            help="Upload the Excel file containing your current timetable"
        )
        
        if uploaded_file is not None:
            st.markdown('<div class="status-success">✅ File uploaded successfully!</div>', unsafe_allow_html=True)
    
    with col2:
        st.markdown("""
        <div class="metric-card">
            <h4>🚀 Optimization Features</h4>
            <ul style="text-align: left; margin: 0;">
                <li>📊 Analyzes empty slots</li>
                <li>📅 Reduces exam span</li>
                <li>🎯 Fills gaps intelligently</li>
                <li>⚡ Maintains constraints</li>
                <li>📈 Shows improvement metrics</li>
            </ul>
        </div>
        """, unsafe_allow_html=True)
    
    if uploaded_file is not None:
        # Process buttons
        col1, col2 = st.columns(2)
        
        with col1:
            if st.button("📊 Analyze Timetable", use_container_width=True):
                with st.spinner("Analyzing timetable..."):
                    # Save uploaded file temporarily
                    temp_file = "temp_upload.xlsx"
                    with open(temp_file, 'wb') as f:
                        f.write(uploaded_file.read())
                    
                    # Read timetable
                    optimizer.read_existing_timetable(temp_file)
                    
                    # Analyze
                    analysis = optimizer.analyze_timetable()
                    
                    # Clean up
                    if os.path.exists(temp_file):
                        os.remove(temp_file)
                    
                    # Display analysis
                    st.markdown("### 📊 Current Timetable Analysis")
                    
                    col1, col2, col3, col4 = st.columns(4)
                    with col1:
                        st.metric("Total Empty Slots", analysis['total_empty_slots'])
                    with col2:
                        st.metric("Current Span (days)", analysis['current_span'])
                    with col3:
                        st.metric("Total Sheets", len(optimizer.timetable_data))
                    with col4:
                        st.metric("OE Exam Days", len(analysis['oe_dates']))
                    
                    # Show empty slots by date
                    if analysis['empty_slots_by_date']:
                        col1, col2 = st.columns(2)
                        
                        with col1:
                            st.markdown("#### Empty Slots by Date")
                            empty_df = pd.DataFrame(
                                list(analysis['empty_slots_by_date'].items()),
                                columns=['Date', 'Empty Slots']
                            ).sort_values('Date')
                            st.dataframe(empty_df, hide_index=True, use_container_width=True)
                        
                        with col2:
                            if analysis['oe_info']:
                                st.markdown("#### Current OE Exam Dates")
                                oe_df = pd.DataFrame(
                                    [(date, ", ".join([f"{ts}" for ts, _ in info])) for date, info in analysis['oe_info'].items()],
                                    columns=['Date', 'Time Slots']
                                ).sort_values('Date')
                                st.dataframe(oe_df, hide_index=True, use_container_width=True)
        
        with col2:
            if st.button("🔧 Optimize Timetable", use_container_width=True, type="primary"):
                if not optimizer.timetable_data:
                    st.error("Please analyze the timetable first!")
                else:
                    with st.spinner("Optimizing timetable..."):
                        # Run optimization
                        results = optimizer.optimize_timetable(holidays)
                        
                        # Display results
                        st.markdown("### ✅ Optimization Results")
                        
                        col1, col2, col3, col4 = st.columns(4)
                        with col1:
                            st.metric(
                                "Days Saved", 
                                results['days_saved'],
                                delta=f"-{results['days_saved']} days"
                            )
                        with col2:
                            st.metric(
                                "Slots Filled",
                                results['slots_filled'],
                                delta=f"+{results['slots_filled']} exams"
                            )
                        with col3:
                            st.metric(
                                "New Span",
                                results['final_analysis']['current_span'],
                                delta=f"{results['final_analysis']['current_span'] - results['initial_analysis']['current_span']} days"
                            )
                        with col4:
                            # Count OE moves in log
                            oe_moves = sum(1 for log in optimizer.optimization_log if "OE exams" in log)
                            st.metric(
                                "OE Slots Moved",
                                oe_moves,
                                delta=f"↑ {oe_moves}"
                            )
                        
                        # Show optimization log
                        if optimizer.optimization_log:
                            with st.expander("📝 Optimization Log", expanded=False):
                                # Separate regular and OE moves
                                regular_moves = [log for log in optimizer.optimization_log if "OE exams" not in log]
                                oe_moves_log = [log for log in optimizer.optimization_log if "OE exams" in log]
                                
                                if regular_moves:
                                    st.markdown("**Regular Subject Moves:**")
                                    for log in regular_moves[-5:]:  # Show last 5
                                        st.write(f"• {log}")
                                
                                if oe_moves_log:
                                    st.markdown("**Open Elective (OE) Moves:**")
                                    for log in oe_moves_log:
                                        st.write(f"• 🎯 {log}")
                        
                        st.session_state.optimization_complete = True
    
    # Download section
    if 'optimization_complete' in st.session_state and st.session_state.optimization_complete:
        st.markdown("---")
        st.markdown("### 📥 Download Optimized Timetable")
        
        col1, col2, col3 = st.columns(3)
        
        with col1:
            excel_data = optimizer.save_optimized_excel()
            st.download_button(
                label="📊 Download Optimized Excel",
                data=excel_data.getvalue(),
                file_name=f"optimized_timetable_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx",
                mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                use_container_width=True
            )
        
        with col2:
            if st.button("📄 Generate PDF", use_container_width=True):
                with st.spinner("Generating PDF..."):
                    pdf_path = "optimized_timetable.pdf"
                    optimizer.generate_optimized_pdf(pdf_path)
                    
                    with open(pdf_path, 'rb') as f:
                        pdf_data = f.read()
                    
                    st.download_button(
                        label="📄 Download PDF",
                        data=pdf_data,
                        file_name=f"optimized_timetable_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf",
                        mime="application/pdf",
                        use_container_width=True
                    )
                    
                    # Clean up
                    if os.path.exists(pdf_path):
                        os.remove(pdf_path)
        
        with col3:
            if st.button("🔄 Reset", use_container_width=True):
                st.session_state.optimizer = TimetableOptimizer()
                st.session_state.optimization_complete = False
                st.rerun()
    
    # Instructions
    st.markdown("---")
    st.markdown("""
    ### 📖 How to Use
    
    1. **Upload**: Upload your existing timetable Excel file
    2. **Analyze**: Click "Analyze Timetable" to see current state
    3. **Optimize**: Click "Optimize Timetable" to fill empty slots
    4. **Download**: Get your optimized timetable in Excel or PDF format
    
    The optimizer will:
    - Identify all empty slots (marked as "---")
    - Move exams from later dates to earlier empty slots
    - Reduce the overall examination span
    - Maintain all constraints and avoid conflicts
    """)

if __name__ == "__main__":
    main()
